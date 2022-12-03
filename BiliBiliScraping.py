from selenium import webdriver
# import selenium
# import webbrowser
# import xlrd
# import pyinputplus as pyip
import threading
import requests
import bs4
import json
import csv
import os
import openpyxl
import re
import logging
from datetime import datetime
import uuid
import multiprocessing
import pprint

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# 搜索URL
searchUrl_original = r'https://api.bilibili.com/x/web-interface/search/all/v2?__refresh__=true&_extra=&context=&page=1&page_size=42&order=&duration=&from_source=&from_spmid=333.337&platform=pc&highlight=1&single_column=0&keyword={}&qv_id=TLHELrINKOkCHqaznD9YglZwIHSiQgx2&ad_resource=5646&source_tag=3'

# 子集搜索URL
episodeUrl_original = r'https://api.bilibili.com/pgc/web/season/section?season_id={}'

# 请求头
headers = {
    'User-Agent': r'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36',
    'Cookie': r'buvid3=F494FDEB-2AF5-264D-F7C3-1515C6E8AC4A81199infoc; b_nut=1669095381; i-wanna-go-back=-1; b_ut=7; b_lsid=14FB6E107_1849DD65C47; _uuid=42561AD4-38B4-D9BD-9210E-B38EB8E5B910682099infoc; buvid_fp=1b5de45027f1f220e6f0da6701ce8227; buvid4=E1FEECDD-3E67-9406-6C56-83EB2E68554E82916-022112213-csQmflR1YMnOgnTW8hOGRw%3D%3D; nostalgia_conf=-1; CURRENT_FNVAL=4048; bsource=search_google; innersign=1; sid=p6wz9elt; CURRENT_QUALITY=16; rpdid=|(J~kkmYJRlk0J\'uYYm|mlm|J'
}

# 爬取片单信息字段
scrapedResultDict = {"名称": '', '描述': '', '标签': '', '地域': '', '海报图': '', '类型': '', '制作信息': '', '演员': '', '上线时间': '',
                     '集数': '', '评分': '', 'UP主': ''}
# 爬取片单结果集
scrapedResultList = []
# 片单子集信息字段
scrapedEpisodeDict = {"标题": '', '长标题': '', '封面': '', '播放地址': ''}
# 爬取节目子集结果集
scrapedEpisodeResult = {}

# 最多允许CORE核数个线程同时运行
semaphore = threading.BoundedSemaphore(multiprocessing.cpu_count())


def __parsePGCSub__(searchVideoName, section):
    """
    :param searchVideoName:
    :param section:
    :return:
    """
    logging.info(f'{searchVideoName}抓取到子集信息 {json.dumps(section)} ')
    episodes = list(section['episodes'])
    scrapedEpisodeList = []

    for episode in episodes:
        scpapedEpisode = scrapedEpisodeDict.copy()
        scpapedEpisode['封面'] = episode['cover']
        scpapedEpisode['播放地址'] = episode['share_url']
        scpapedEpisode['标题'] = episode['title']
        scpapedEpisode['长标题'] = episode['long_title']
        scrapedEpisodeList.append(scpapedEpisode)

    # EXCEL sheet 中不能有如下特殊字符
    searchVideoName = re.sub(r'\?|？|：|\\|/|\*|\:', '', searchVideoName)
    scrapedEpisodeResult[searchVideoName] = scrapedEpisodeList


class ScrapingThread(threading.Thread):

    def __init__(self, searchVideoName, semaphore):
        """
        :param searchVideoName: 搜索视频名称
        :param semaphore:
        """
        super(ScrapingThread, self).__init__()
        self.searchVideoName = searchVideoName
        self.semaphore = semaphore
        self.flag = uuid.uuid1()

    def __parsePGC__(searchVideoName, item):
        """
        解析PUC视频信息
        :param searchVideoName: 视频的名称
        :param item:            爬取到的相关数据
        :return:
        """
        logging.info(f'{searchVideoName}抓取到信息 {json.dumps(item)}')
        videoDetailList = item['data']
        videoDetail = videoDetailList[0]
        scrapedResultDetail = scrapedResultDict.copy()

        scrapedResultDetail['名称'] = searchVideoName

        scrapedResultDetail['标签'] = videoDetail['styles'].replace(",", "，")
        scrapedResultDetail['地域'] = videoDetail['areas']
        scrapedResultDetail['海报图'] = videoDetail['cover']
        scrapedResultDetail['类型'] = videoDetail['season_type_name']

        staff = str(videoDetail['staff'])
        staff = removeHtmlTag(staff)
        scrapedResultDetail['制作信息'] = staff

        scrapedResultDetail['演员'] = videoDetail['cv']

        dt_now = datetime.fromtimestamp(int(videoDetail['pubtime']))
        pubTime = dt_now.strftime("%Y.%m.%d")
        scrapedResultDetail['上线时间'] = str(pubTime)

        scrapedResultDetail['集数'] = videoDetail['index_show']

        media_score = videoDetail['media_score']
        score = media_score['score']

        scrapedResultDetail['评分'] = score

        gotoUrl = videoDetail['goto_url']

        res = requests.get(gotoUrl).text

        mediaCover = bs4.BeautifulSoup(res, features="html.parser").select('.media-cover')
        detailPageUrl = mediaCover[0].attrs['href']
        detailPageUrl = "https:" + detailPageUrl

        detailPage = requests.get(detailPageUrl).text
        detailJson = json.loads(re.search(r"window\.__INITIAL_STATE__=(.*?);", detailPage).group(1))
        mediaInfo = detailJson['mediaInfo']
        seasonId = mediaInfo['season_id']
        evaluate = mediaInfo['evaluate']
        scrapedResultDetail['描述'] = evaluate

        # 节目子集信息获取接口
        episodeUrl = episodeUrl_original.format(seasonId)

        episodesInfo = requests.get(episodeUrl).text
        result = json.loads(episodesInfo)['result']

        if 'main_section' in result:
            section = result['main_section']
            __parsePGCSub__(searchVideoName, section)
        elif 'section' in result:
            section = result['section']
            __parsePGCSub__(searchVideoName, section)

            # writeExcelFile(resultFile, searchVideoName, list(scrapedEpisodeDict.keys()), scrapedEpisodeList)

        scrapedResultList.append(scrapedResultDetail)


    def __parseUGC__(searchVideoName, item):

        logging.info(f'{searchVideoName} 抓取到信息 {json.dumps(item)} ')

        videoDetailList = item['data']
        videoDetail = videoDetailList[0]

        scrapedResultDetail = scrapedResultDict.copy()
        scrapedResultDetail['名称'] = searchVideoName
        desc = str(videoDetail['description']).replace(",", "，")
        desc = removeHtmlTag(desc)
        scrapedResultDetail['描述'] = desc
        scrapedResultDetail['海报图'] = str('https:') + videoDetail['pic']
        scrapedResultDetail['标签'] = videoDetail['tag'].replace(",", "/")
        scrapedResultDetail['类型'] = videoDetail['typename']

        dt_now = datetime.fromtimestamp(int(videoDetail['pubdate']))
        pubTime = dt_now.strftime("%Y.%m.%d")
        scrapedResultDetail['上线时间'] = str(pubTime)

        scrapedResultDetail['UP主'] = videoDetail['author']

        scrapedResultList.append(scrapedResultDetail)

    def run(self):

        self.semaphore.acquire()  # 加锁

        searchVideoName = self.searchVideoName
        logging.info(f"开始爬取: {searchVideoName}")

        searchUrl = searchUrl_original.format(searchVideoName)

        try:
            res = requests.get(searchUrl, headers=headers)
            # The raise_for_status() method is a good way to ensure that a program halts if a bad download occurs.
            res.raise_for_status()
        except Exception as ex:
            logging.error(f"{searchVideoName} 爬取出错了 {ex}")
            return

        noStarchSoup = json.loads(bs4.BeautifulSoup(res.text, 'html.parser').text)
        noStarchSoupData = noStarchSoup['data']
        noStarchSoupDataResult = list(noStarchSoupData['result'])

        for item in noStarchSoupDataResult:
            scrapedResultDetail = scrapedResultDict.copy()
            scrapedResultDetail['名称'] = searchVideoName
            if item['result_type'] == 'media_bangumi' and len(item['data']) > 0:
                self.__parsePGC__(searchVideoName, item)
                break
            elif item['result_type'] == 'media_ft' and len(item['data']) > 0:
                self.__parsePGC__(searchVideoName, item)
                break
            elif item['result_type'] == 'video' and len(item['data']) > 0:
                self.__parseUGC__(searchVideoName, item)
                break

        self.semaphore.release()  # 释放


def removeHtmlTag(content):
    """
    :param content: raw data needs to eliminate the html related flags
    :return:
    """
    content = re.sub(r'<em .*\">', "", content)
    content = re.sub(r'</em>', "", content)
    return content


def writeExcelFile(file, sheetName, row_headers, rows):
    """
    :param file:
    :param sheetName:
    :param row_headers:
    :param rows:
    :return:
    """
    if os.path.isfile(file):
        wb = openpyxl.load_workbook(file)
    else:
        wb = openpyxl.Workbook()

    if sheetName in wb.sheetnames:
        sheet = wb[sheetName]
    else:
        sheet = wb.create_sheet(title=sheetName)

    for index, head in enumerate(row_headers):
        sheet.cell(row=1, column=index + 1).value = head
        for rowNum in range(2, len(rows) + 2):
            value = rows[rowNum - 2][head]
            if value is not None:
                sheet.cell(row=rowNum, column=index + 1).value = value

    wb.save(file)


def writeCsvFile(file, header, rows):
    """
    :param file: CSV file saved the scraping data
    :param header: result headers
    :param rows: data
    :return:
    """

    """
    DictReader and DictWriter CSV Objects
    For CSV files that contain header rows, it’s often more convenient to
    work with the DictReader and DictWriter objects, rather than the reader and writer objects.
    On Windows, you’ll also need to pass a blank string for the open() function’s newline keyword argument
    if you forget to set the newline argument, the rows in output.csv will be double-spaced
    """
    with open(file, 'w', newline='') as csvFile:
        """
        The delimiter and lineterminator Keyword Arguments
        separate cells with a tab character instead of a comma and you want the rows to be double-spaced.
        csv.writer(csvFile, delimiter='\t', lineterminator='\n\n')
        """

        writer = csv.DictWriter(csvFile, header)
        writer.writeheader()
        writer.writerows(rows)


if __name__ == '__main__':

    excel_path = input("请输入excel(xlsx):")
    while not (os.path.isfile(excel_path) and os.path.basename(excel_path).endswith("xlsx")):
        os.path.basename(excel_path)
        print("输入的EXCEL不正确，请重新输入")
        excel_path = input("请输入excel(xlsx):")

    # 保存路径
    deskTopPath = os.path.join(os.path.expanduser("~"), 'Desktop')
    resultFile = os.path.join(deskTopPath, 'result.xlsx')

    browser = webdriver.Firefox()
    browser.get("https://www.bilibili.com/")

    cookieStr = ''
    cookies = browser.get_cookies()
    for cookie in browser.get_cookies():
        cookieStr += (cookie['name'] + "=" + cookie['value'] + ";")

    headers['Cookie'] = cookieStr
    browser.close()

    try:
        workbook = openpyxl.load_workbook(excel_path)
    except Exception as ex:
        logging.error(f'打开{excel_path} 失败 {ex}')

    sheet = workbook.active
    cells = list(sheet.columns)[0]

    for cellObj in cells:
        searchVideoName = cellObj.value
        # searchVideoName = "2021最美的夜 bilibili晚会"
        # p = re.compile(r'《(.*)》')
        # mo = p.search(str(searchVideoName))
        # if mo is not None:
        #     searchVideoName = mo.group(1)

        t = ScrapingThread(searchVideoName, semaphore)
        t.start()

    logging.info(f'保存结果文件: {resultFile}')
    writeExcelFile(resultFile, 'Sheet', list(scrapedResultDict.keys()), scrapedResultList)

    for movieName in scrapedEpisodeResult.keys():
        writeExcelFile(resultFile, movieName, list(scrapedEpisodeDict.keys()), scrapedEpisodeResult[movieName])

    logging.info('本次任务完成')
